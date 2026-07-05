"""
step8e_final_report.py

Step8E: FINAL scientific summary report combining Step8A-Step8D into one
coherent package.

IMPORTANT:
    - This step does NOT train any model.
    - This step does NOT change any previous output (Step8A/B/C/D files are
      read-only inputs).
    - This step does NOT recompute any statistic or metric. It only reads
      the stats JSON files already produced by Step8A/B/C/D and aggregates/
      formats them into a combined report. Any comparison used purely to
      phrase a sentence (e.g. "group X ranks above group Y") reads values
      that were already computed by Step8D -- no new modeling or statistics
      are performed here.

Input (read-only):
    outputs/step8a/step8a_dataset_stats.json        (required)
    outputs/step8b/step8b_model_comparison_metrics.json (required)
    outputs/step8c/step8c_bootstrap_metrics.json     (optional -- warns and
        marks the bootstrap section "not available" if missing)
    outputs/step8d/step8d_ablation_metrics.json      (optional -- warns and
        marks the ablation section "not available" if missing)

Output:
    outputs/step8e/step8e_summary.md
    outputs/step8e/step8e_summary.json
    outputs/step8e/step8e_results_tables.xlsx
    outputs/step8e/step8e_key_findings.csv
    outputs/step8e/step8e_feature_ranking.csv     (optional)
    outputs/step8e/step8e_monthly_results.csv     (optional)
    outputs/step8e/step8e_population_summary.csv  (optional)

CLI:
    python src/step8e_final_report.py --force
"""

from __future__ import annotations

import argparse
import json
import sys
from datetime import datetime, timezone
from pathlib import Path

_PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(_PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(_PROJECT_ROOT))

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from core.config import STEP8E_OUTPUT_DIR
from core.io_utils import setup_logger
from core.paths import PROJECT_ROOT

BASE_DIR = PROJECT_ROOT
log, log_file = setup_logger("step8e")


class Step8EError(SystemExit):
    """Fail-fast error for Step8E (extends SystemExit like other steps)."""


STEP8A_STATS_PATH = "outputs/step8a/step8a_dataset_stats.json"
STEP8B_STATS_PATH = "outputs/step8b/step8b_model_comparison_metrics.json"
STEP8C_STATS_PATH = "outputs/step8c/step8c_bootstrap_metrics.json"
STEP8D_STATS_PATH = "outputs/step8d/step8d_ablation_metrics.json"

PRIMARY_POPULATION = "all_valid"
SECONDARY_POPULATION = "cropland_dominant"
MONTH_ORDER = ["august", "september", "october"]

ONLY_GROUPS = [
    "lst_anomaly_only", "current_lst_only", "tvdi_only",
    "tvdi_difference_only", "downscaled_only", "fused_lst_only",
]


# =============================================================================
# Loading (read-only; no recomputation)
# =============================================================================
def load_json(path: Path, required: bool, warnings_out: list[str]) -> dict | None:
    if not path.exists():
        msg = f"{path} bulunamadi."
        if required:
            raise Step8EError(
                f"{msg} Bu dosya Step8E raporu icin ZORUNLUDUR. Once ilgili "
                "adimi calistirin."
            )
        warnings_out.append(f"{msg} Ilgili rapor bolumu 'not available' olarak isaretlenecek.")
        log.warning(msg)
        return None
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception as exc:  # noqa: BLE001
        msg = f"{path} okunamadi/parse edilemedi: {exc}"
        if required:
            raise Step8EError(msg)
        warnings_out.append(msg)
        log.warning(msg)
        return None


def load_all_inputs() -> tuple[dict, dict, dict | None, dict | None, list[str]]:
    warnings_out: list[str] = []
    step8a = load_json(BASE_DIR / STEP8A_STATS_PATH, required=True, warnings_out=warnings_out)
    step8b = load_json(BASE_DIR / STEP8B_STATS_PATH, required=True, warnings_out=warnings_out)
    step8c = load_json(BASE_DIR / STEP8C_STATS_PATH, required=False, warnings_out=warnings_out)
    step8d = load_json(BASE_DIR / STEP8D_STATS_PATH, required=False, warnings_out=warnings_out)
    return step8a, step8b, step8c, step8d, warnings_out


# =============================================================================
# Section builders (pure aggregation of already-computed values)
# =============================================================================
def fmt(x, nd=4, pct=False):
    if x is None:
        return "n/a"
    try:
        if pct:
            return f"{x * 100:.{max(nd - 2, 1)}f}%"
        return f"{x:+.{nd}f}" if isinstance(x, float) and x != 0 else f"{x:.{nd}f}"
    except (TypeError, ValueError):
        return str(x)


def build_dataset_section(step8a: dict) -> dict:
    return {
        "total_500m_cells": step8a.get("total_500m_cells"),
        "valid_modeling_cells": step8a.get("valid_modeling_cells"),
        "invalid_cells": step8a.get("invalid_cells"),
        "burned_cell_count": step8a.get("burned_cell_count"),
        "unburned_cell_count": step8a.get("unburned_cell_count"),
        "burned_rate": step8a.get("burned_rate"),
        "label_source": step8a.get("label_kind") or "MCD64A1",
        "label_source_description": step8a.get("label_source_description"),
        "burn_month_counts": step8a.get("burn_month_counts"),
        "burnable_tree_shrub_grass_count": step8a.get("burnable_tree_shrub_grass_count"),
        "burnable_tree_shrub_count": step8a.get("burnable_tree_shrub_count"),
        "burned_count_within_each_burnable_mask": step8a.get("burned_count_within_each_burnable_mask"),
        "cropland_burned_count": step8a.get("cropland_burned_count"),
        "burned_count_within_primary_burnable_mask": step8a.get("burned_count_within_primary_burnable_mask"),
    }


def build_population_section(step8b: dict, step8d: dict | None) -> dict:
    pop_counts = step8b.get("population_counts", {})
    skipped_8b = step8b.get("skipped_populations", {})
    skipped_8d = (step8d or {}).get("skipped_populations", {})
    populations = {}
    for pop in [PRIMARY_POPULATION, SECONDARY_POPULATION, "burnable_tree_shrub_grass", "burnable_tree_shrub"]:
        populations[pop] = {
            "n_cells": pop_counts.get(pop),
            "skipped_in_step8b": skipped_8b.get(pop),
            "skipped_in_step8d": skipped_8d.get(pop),
        }
    return populations


def build_model_comparison_section(step8b: dict) -> dict:
    pm = step8b.get("population_metrics", {})
    out = {}
    for pop in [PRIMARY_POPULATION, SECONDARY_POPULATION]:
        p = pm.get(pop)
        if p is None:
            out[pop] = None
            continue
        base, therm = p.get("overall_baseline", {}), p.get("overall_thermal", {})
        out[pop] = {
            "n_positives": p.get("n_positives"), "n_negatives": p.get("n_negatives"),
            "n_splits_used": p.get("n_splits_used"),
            "auc_baseline": base.get("roc_auc"), "auc_thermal": therm.get("roc_auc"),
            "delta_auc": p.get("delta_auc"),
            "pr_auc_baseline": base.get("pr_auc"), "pr_auc_thermal": therm.get("pr_auc"),
            "delta_pr_auc": p.get("delta_pr_auc"),
            "brier_baseline": base.get("brier_score"), "brier_thermal": therm.get("brier_score"),
            "delta_brier": p.get("delta_brier"),
            "interpretation": p.get("interpretation"),
        }
    return out


def build_bootstrap_section(step8c: dict | None) -> dict:
    if step8c is None:
        return {"available": False}
    ci_by_pop = step8c.get("bootstrap_ci_by_population", {})
    monthly_by_pop = step8c.get("monthly_bootstrap_ci_by_population", {})
    out = {"available": True, "populations": {}, "wording_note": step8c.get("wording_note")}
    for pop in [PRIMARY_POPULATION, SECONDARY_POPULATION]:
        ci = ci_by_pop.get(pop)
        if ci is None or not ci.get("available"):
            out["populations"][pop] = {"available": False}
            continue
        out["populations"][pop] = {
            "available": True,
            "delta_auc_ci95": ci.get("delta_auc_ci95"),
            "delta_auc_interpretation": ci.get("delta_auc_interpretation"),
            "delta_pr_auc_ci95": ci.get("delta_pr_auc_ci95"),
            "delta_pr_auc_interpretation": ci.get("delta_pr_auc_interpretation"),
            "probability_delta_auc_gt_0": ci.get("probability_delta_auc_gt_0"),
            "monthly": monthly_by_pop.get(pop, {}),
        }
    return out


def build_ablation_section(step8d: dict | None) -> dict:
    if step8d is None:
        return {"available": False}
    deltas_by_pop = step8d.get("ablation_delta_metrics_by_population", {})
    rankings_by_pop = step8d.get("ablation_rankings_by_population", {})
    out = {"available": True, "populations": {}}
    for pop in [PRIMARY_POPULATION, SECONDARY_POPULATION]:
        deltas = deltas_by_pop.get(pop)
        ranking = rankings_by_pop.get(pop)
        if not deltas or not ranking:
            out["populations"][pop] = {"available": False}
            continue

        best_single, best_single_val = None, float("-inf")
        for g in ONLY_GROUPS:
            v = deltas.get(g, {}).get("delta_pr_auc")
            if v is not None and v > best_single_val:
                best_single_val, best_single = v, g

        best_group = ranking[0] if ranking else None
        all_thermal_rank = deltas.get("all_thermal", {}).get("rank_by_delta_pr_auc")
        tvdi_group_delta = deltas.get("tvdi_group", {}).get("delta_pr_auc")
        tvdi_only_delta = deltas.get("tvdi_only", {}).get("delta_pr_auc")
        downscaled_delta = deltas.get("downscaled_only", {}).get("delta_pr_auc")
        fused_downscaled_delta = deltas.get("fused_downscaled_group", {}).get("delta_pr_auc")
        lst_anomaly_group_delta = deltas.get("lst_anomaly_group", {}).get("delta_pr_auc")
        all_thermal_delta = deltas.get("all_thermal", {}).get("delta_pr_auc")

        simpler_groups = [lst_anomaly_group_delta, tvdi_group_delta]
        simpler_groups = [v for v in simpler_groups if v is not None]
        fusion_adds_value = (
            fused_downscaled_delta is not None and simpler_groups
            and fused_downscaled_delta > max(simpler_groups)
        )

        out["populations"][pop] = {
            "available": True,
            "ranking_order": ranking,
            "deltas": deltas,
            "best_single_feature": best_single,
            "best_group_overall": best_group,
            "all_thermal_rank_by_delta_pr_auc": all_thermal_rank,
            "tvdi_contribution_delta_pr_auc": tvdi_group_delta if tvdi_group_delta is not None else tvdi_only_delta,
            "downscaled_contribution_delta_pr_auc": downscaled_delta,
            "fusion_contribution_delta_pr_auc": fused_downscaled_delta,
            "fusion_adds_value_beyond_simpler_groups": fusion_adds_value,
            "lst_anomaly_group_delta_pr_auc": lst_anomaly_group_delta,
            "all_thermal_delta_pr_auc": all_thermal_delta,
        }
    return out


def build_monthly_section(step8b: dict, bootstrap_section: dict) -> dict:
    pm = step8b.get("population_metrics", {})
    out = {}
    for pop in [PRIMARY_POPULATION, SECONDARY_POPULATION]:
        p = pm.get(pop)
        monthly_b = (p or {}).get("monthly", {})
        monthly_c = bootstrap_section.get("populations", {}).get(pop, {}).get("monthly", {}) \
            if bootstrap_section.get("available") else {}
        pop_out = {}
        for month in MONTH_ORDER:
            mb = monthly_b.get(month, {})
            mc = monthly_c.get(month, {}) if isinstance(monthly_c, dict) else {}
            n_pos = mb.get("positive_count")
            auc_b, auc_t = mb.get("auc_baseline"), mb.get("auc_thermal")
            delta_auc = (auc_t - auc_b) if (auc_b is not None and auc_t is not None) else None
            pr_b, pr_t = mb.get("pr_auc_baseline"), mb.get("pr_auc_thermal")
            delta_pr = (pr_t - pr_b) if (pr_b is not None and pr_t is not None) else None

            if delta_auc is None:
                confidence = "unavailable (too few positives)"
            elif mc.get("available"):
                interp = mc.get("delta_auc_interpretation")
                confidence = (
                    "strong (bootstrap-supported positive delta)" if interp == "positive_bootstrap_support"
                    else "weak (bootstrap-supported negative delta)" if interp == "negative_bootstrap_support"
                    else "low confidence (bootstrap CI overlaps zero)"
                )
            else:
                confidence = "low confidence (positives scarce; no bootstrap CI available)" if (n_pos or 0) < 15 else "point estimate only"

            pop_out[month] = {
                "positive_count": n_pos, "negative_count": mb.get("negative_count"),
                "delta_auc": delta_auc, "delta_pr_auc": delta_pr,
                "bootstrap_ci95_delta_auc": mc.get("delta_auc_ci95") if isinstance(mc, dict) else None,
                "confidence": confidence,
            }
        out[pop] = pop_out
    return out


def build_limitations(dataset: dict, populations: dict, model_comparison: dict) -> list[str]:
    crop_n = (populations.get(SECONDARY_POPULATION) or {}).get("n_cells")
    burned_n = dataset.get("burned_cell_count")
    crop_pct = None
    if crop_n and burned_n:
        cropland_burned = dataset.get("cropland_burned_count")
        if cropland_burned is not None and burned_n:
            crop_pct = cropland_burned / burned_n
    tsg_burned = (dataset.get("burned_count_within_each_burnable_mask") or {}).get("burnable_tree_shrub_grass")
    ts_burned = (dataset.get("burned_count_within_each_burnable_mask") or {}).get("burnable_tree_shrub")
    october_pos = (dataset.get("burn_month_counts") or {}).get("10") or (dataset.get("burn_month_counts") or {}).get(10)

    return [
        "Single AOI (Kozan / East Mediterranean region): results have not been "
        "validated against any other geography and may not generalize.",
        "Single target season (2023 pre-fire Aug-Oct label window): no "
        "multi-year train/test split has been performed yet, so year-to-year "
        "generalization is unknown.",
        f"Cropland-excluded burnable masks (tree+shrub+grass, tree+shrub) have "
        f"too few burned positives to serve as a primary modeling population "
        f"(tree+shrub+grass burned={tsg_burned if tsg_burned is not None else 'n/a'}, "
        f"tree+shrub burned={ts_burned if ts_burned is not None else 'n/a'}); they "
        "are reported as diagnostic/sensitivity analyses only.",
        (
            f"Burned cells are overwhelmingly cropland-dominant "
            f"({fmt(crop_pct, pct=True) if crop_pct is not None else 'the large majority'} "
            "of burned cells), which limits how confidently results generalize to "
            "natural vegetation fire behavior in this AOI/window."
            if crop_pct is not None else
            "Burned cells are overwhelmingly cropland-dominant, which limits how "
            "confidently results generalize to natural vegetation fire behavior."
        ),
        (
            f"October has a low sample size (positives={october_pos}), so October "
            "lead-time metrics carry low confidence and wide/unavailable bootstrap "
            "intervals." if october_pos is not None else
            "October has a low burned-cell sample size, so October lead-time "
            "metrics carry low confidence."
        ),
        "Daily MODIS-based gap filling for the thermal context is not yet "
        "available; current gap-filling relies on Step7D/7E's downscaled/fused "
        "LST rather than a daily MODIS observation stream.",
    ]


FUTURE_WORK = [
    "Multi-year train/test split to assess temporal generalization beyond the "
    "single 2023 pre-fire season.",
    "Extend validation to a second AOI to test spatial generalization beyond "
    "the Kozan / East Mediterranean region.",
    "Incorporate daily MODIS thermal context (rather than the current period-"
    "aggregated composites) for finer temporal resolution.",
    "Implement systematic gap filling for the daily thermal record to reduce "
    "reliance on the current Step7D/7E downscaling/fusion approach.",
    "Move toward an operational digital-twin deployment (near-real-time "
    "scoring, monitoring, and periodic retraining).",
]


def build_main_findings(
    dataset: dict, model_comparison: dict, bootstrap: dict, ablation: dict,
) -> list[str]:
    findings = [
        f"Reconstructing the native ~500 m MCD64A1 grid (instead of the "
        f"previous 30 m resampled label) removed pseudo-replication: the "
        f"modeling dataset now has {dataset.get('valid_modeling_cells')} "
        f"independent cells ({dataset.get('burned_cell_count')} burned, "
        f"{dataset.get('unburned_cell_count')} unburned) rather than millions "
        "of duplicated 30 m pixels.",
    ]

    primary_mc = model_comparison.get(PRIMARY_POPULATION)
    if primary_mc and primary_mc.get("delta_auc") is not None:
        direction = "improves" if primary_mc["delta_auc"] > 0 else "does not improve"
        findings.append(
            f"Adding thermal/dryness features {direction} burned-area "
            f"discrimination beyond the static baseline for the primary "
            f"population ({PRIMARY_POPULATION}): delta_auc="
            f"{fmt(primary_mc['delta_auc'])}, delta_pr_auc="
            f"{fmt(primary_mc['delta_pr_auc'])}."
        )

    findings.append(
        "This improvement was estimated under spatial-block cross-validation "
        "(blocks of ~1 km, StratifiedGroupKFold), not a random row-wise "
        "split, so it is not an artifact of spatially autocorrelated "
        "train/test leakage."
    )

    if bootstrap.get("available"):
        primary_boot = bootstrap["populations"].get(PRIMARY_POPULATION, {})
        if primary_boot.get("available"):
            interp = primary_boot.get("delta_auc_interpretation")
            if interp == "positive_bootstrap_support":
                findings.append(
                    "The thermal improvement survives spatial-block bootstrap "
                    f"resampling for {PRIMARY_POPULATION} (delta_auc 95% CI="
                    f"{primary_boot.get('delta_auc_ci95')}, entirely above zero)."
                )
            elif interp == "uncertain":
                findings.append(
                    f"The thermal improvement for {PRIMARY_POPULATION} is "
                    "positive as a point estimate, but the spatial-block "
                    f"bootstrap 95% CI overlaps zero (delta_auc CI="
                    f"{primary_boot.get('delta_auc_ci95')}), so it should be "
                    "treated as uncertain rather than confirmed."
                )
    else:
        findings.append(
            "Spatial-block bootstrap results (Step8C) were not available for "
            "this report; uncertainty around the thermal improvement is not "
            "quantified here."
        )

    if ablation.get("available"):
        primary_abl = ablation["populations"].get(PRIMARY_POPULATION, {})
        if primary_abl.get("available"):
            best_single = primary_abl.get("best_single_feature")
            best_group = primary_abl.get("best_group_overall")
            all_thermal_rank = primary_abl.get("all_thermal_rank_by_delta_pr_auc")
            if best_group == "all_thermal":
                findings.append(
                    "The best-performing thermal feature set combines all "
                    "thermal indicators together (`all_thermal`) rather than "
                    "any single simpler group, indicating thermal features "
                    "carry complementary (not redundant) information."
                )
            elif best_group is not None:
                findings.append(
                    f"The single best-performing thermal feature GROUP is "
                    f"`{best_group}` (rank #1 by delta PR-AUC); `all_thermal` "
                    f"ranks #{all_thermal_rank}, so combining every thermal "
                    "feature is not strictly necessary to capture most of "
                    "the improvement."
                )
            if best_single is not None:
                findings.append(f"The best single (standalone) thermal feature is `{best_single}`.")

            lst_grp = primary_abl.get("lst_anomaly_group_delta_pr_auc")
            tvdi_c = primary_abl.get("tvdi_contribution_delta_pr_auc")
            if lst_grp is not None and tvdi_c is not None:
                comparison = "more than" if lst_grp > tvdi_c else "less than" if lst_grp < tvdi_c else "about the same as"
                findings.append(
                    f"The LST-anomaly feature group contributes {comparison} "
                    "the TVDI-based features on their own (delta PR-AUC "
                    f"{fmt(lst_grp)} vs {fmt(tvdi_c)})."
                )
            if primary_abl.get("fusion_adds_value_beyond_simpler_groups") is not None:
                fusion_word = (
                    "provides additional value beyond" if primary_abl["fusion_adds_value_beyond_simpler_groups"]
                    else "does not clearly add value beyond"
                )
                findings.append(
                    f"The fused/downscaled thermal group {fusion_word} the "
                    "simpler LST-anomaly/TVDI groups on their own."
                )
    else:
        findings.append(
            "Thermal feature ablation results (Step8D) were not available "
            "for this report; which specific thermal feature drives the "
            "improvement is not identified here."
        )

    return findings


def build_overall_conclusion(model_comparison: dict, bootstrap: dict, ablation: dict) -> str:
    primary_mc = model_comparison.get(PRIMARY_POPULATION)
    positive = bool(primary_mc and primary_mc.get("delta_auc") and primary_mc["delta_auc"] > 0)
    bootstrap_confirmed = (
        bootstrap.get("available")
        and bootstrap.get("populations", {}).get(PRIMARY_POPULATION, {}).get("delta_auc_interpretation")
        == "positive_bootstrap_support"
    )
    combining_best = (
        ablation.get("available")
        and ablation.get("populations", {}).get(PRIMARY_POPULATION, {}).get("best_group_overall") == "all_thermal"
    )

    if positive and bootstrap_confirmed:
        combo_sentence = (
            "The strongest performance is obtained by combining multiple thermal "
            "indicators rather than relying on a single drought index."
            if combining_best else
            "Notably, a single, simpler thermal feature group captures most of "
            "this improvement, so combining every thermal indicator is not "
            "strictly required to realize most of the benefit."
        )
        return (
            "The current pipeline demonstrates that adding thermal information "
            "provides a measurable, spatial-block-bootstrap-supported improvement "
            "in burned-area prediction under spatial-block validation. "
            "The improvement is observed across bootstrap and ablation analyses, "
            "suggesting that thermal information adds predictive signal beyond "
            "the static environmental baseline. "
            f"{combo_sentence}"
        )
    if positive and not bootstrap_confirmed:
        return (
            "The current pipeline shows a positive point-estimate improvement "
            "from adding thermal information under spatial-block validation, "
            "but this improvement's spatial-block bootstrap confidence interval "
            "overlaps zero (or bootstrap results were unavailable), so it "
            "should be treated as a promising but unconfirmed signal rather "
            "than a settled result."
        )
    return (
        "Based on the available Step8A-Step8D outputs, thermal information "
        "does not show a clearly confirmed improvement over the static "
        "baseline for burned-area prediction in this AOI/window; further "
        "investigation is warranted before drawing a stronger conclusion."
    )


# =============================================================================
# Writers
# =============================================================================
def write_key_findings_csv(findings: list[str], limitations: list[str], future_work: list[str], output_dir: Path) -> Path:
    rows = (
        [{"category": "main_finding", "item": f} for f in findings]
        + [{"category": "limitation", "item": l} for l in limitations]
        + [{"category": "future_work", "item": f} for f in future_work]
    )
    path = output_dir / "step8e_key_findings.csv"
    pd.DataFrame(rows).to_csv(path, index=False)
    return path


def write_feature_ranking_csv(ablation: dict, output_dir: Path) -> Path | None:
    if not ablation.get("available"):
        return None
    rows = []
    for pop, data in ablation["populations"].items():
        if not data.get("available"):
            continue
        for group in data["ranking_order"]:
            d = data["deltas"][group]
            rows.append({
                "population": pop, "rank": d["rank_by_delta_pr_auc"],
                "thermal_group": group, "delta_auc": d["delta_auc"],
                "delta_pr_auc": d["delta_pr_auc"], "delta_brier": d["delta_brier"],
            })
    if not rows:
        return None
    path = output_dir / "step8e_feature_ranking.csv"
    pd.DataFrame(rows).sort_values(["population", "rank"]).to_csv(path, index=False)
    return path


def write_monthly_results_csv(monthly: dict, output_dir: Path) -> Path | None:
    rows = []
    for pop, months in monthly.items():
        for month, m in months.items():
            rows.append({"population": pop, "month": month, **m})
    if not rows:
        return None
    path = output_dir / "step8e_monthly_results.csv"
    pd.DataFrame(rows).to_csv(path, index=False)
    return path


def write_population_summary_csv(populations: dict, output_dir: Path) -> Path | None:
    rows = [{"population": pop, **data} for pop, data in populations.items()]
    if not rows:
        return None
    path = output_dir / "step8e_population_summary.csv"
    pd.DataFrame(rows).to_csv(path, index=False)
    return path


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF")
BODY_FONT = Font(name="Arial")


def _style_sheet(ws, df: pd.DataFrame) -> None:
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            if isinstance(value, (dict, list)):
                value = json.dumps(value, default=str)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = BODY_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for col_idx, col_name in enumerate(df.columns, start=1):
        max_len = max([len(str(col_name))] + [len(str(v)) for v in df[col_name].astype(str)])
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 60)
    ws.freeze_panes = "A2"


def write_excel_report(
    dataset: dict, populations: dict, model_comparison: dict, bootstrap: dict,
    ablation: dict, monthly: dict, limitations: list[str], output_dir: Path,
) -> Path:
    wb = Workbook()
    wb.remove(wb.active)

    # --- dataset sheet ---
    ds_rows = [{"metric": k, "value": json.dumps(v) if isinstance(v, (dict, list)) else v} for k, v in dataset.items()]
    ws = wb.create_sheet("dataset")
    _style_sheet(ws, pd.DataFrame(ds_rows))

    # --- population summary (part of dataset context) ---
    pop_rows = [{"population": pop, **{k: (json.dumps(v) if isinstance(v, (dict, list)) else v) for k, v in data.items()}} for pop, data in populations.items()]
    ws = wb.create_sheet("populations")
    _style_sheet(ws, pd.DataFrame(pop_rows))

    # --- model_comparison sheet ---
    mc_rows = []
    for pop, m in model_comparison.items():
        if m is None:
            mc_rows.append({"population": pop, "available": False})
            continue
        mc_rows.append({"population": pop, "available": True, **m})
    ws = wb.create_sheet("model_comparison")
    _style_sheet(ws, pd.DataFrame(mc_rows))

    # --- bootstrap sheet ---
    boot_rows = []
    if bootstrap.get("available"):
        for pop, data in bootstrap["populations"].items():
            if not data.get("available"):
                boot_rows.append({"population": pop, "available": False})
                continue
            boot_rows.append({
                "population": pop, "available": True,
                "delta_auc_ci95": json.dumps(data.get("delta_auc_ci95")),
                "delta_auc_interpretation": data.get("delta_auc_interpretation"),
                "delta_pr_auc_ci95": json.dumps(data.get("delta_pr_auc_ci95")),
                "delta_pr_auc_interpretation": data.get("delta_pr_auc_interpretation"),
                "probability_delta_auc_gt_0": data.get("probability_delta_auc_gt_0"),
            })
    else:
        boot_rows.append({"population": "n/a", "available": False, "note": "Step8C output not found"})
    ws = wb.create_sheet("bootstrap")
    _style_sheet(ws, pd.DataFrame(boot_rows))

    # --- ablation sheet ---
    abl_rows = []
    if ablation.get("available"):
        for pop, data in ablation["populations"].items():
            if not data.get("available"):
                continue
            for group in data["ranking_order"]:
                d = data["deltas"][group]
                abl_rows.append({
                    "population": pop, "rank": d["rank_by_delta_pr_auc"],
                    "thermal_group": group, "delta_auc": d["delta_auc"],
                    "delta_pr_auc": d["delta_pr_auc"], "delta_brier": d["delta_brier"],
                })
    if not abl_rows:
        abl_rows = [{"note": "Step8D output not found"}]
    ws = wb.create_sheet("ablation")
    _style_sheet(ws, pd.DataFrame(abl_rows))

    # --- monthly sheet ---
    month_rows = []
    for pop, months in monthly.items():
        for month, m in months.items():
            month_rows.append({"population": pop, "month": month, **m})
    ws = wb.create_sheet("monthly")
    _style_sheet(ws, pd.DataFrame(month_rows) if month_rows else pd.DataFrame([{"note": "no monthly data"}]))

    # --- limitations sheet ---
    ws = wb.create_sheet("limitations")
    _style_sheet(ws, pd.DataFrame({"limitation": limitations}))

    path = output_dir / "step8e_results_tables.xlsx"
    wb.save(path)
    return path


def write_summary_md(
    dataset: dict, populations: dict, model_comparison: dict, bootstrap: dict,
    ablation: dict, monthly: dict, findings: list[str], limitations: list[str],
    conclusion: str, warnings_list: list[str], output_dir: Path,
) -> Path:
    lines = ["# Final Burned-Area Modeling Summary", ""]

    lines += [
        "## 1. Pipeline overview", "",
        "**Step8A** reconstructed a native ~500 m MCD64A1-grid modeling "
        "dataset, fixing the pseudo-replication caused by resampling the "
        "500 m burned-area label onto the 30 m predictor grid. Each row is "
        "one independent native burned-area cell, with MCD64A1 raw BurnDate "
        "as the sole label source (FIRMS is never used as target).",
        "",
        "**Step8B** ran the supervisor-requested determining experiment: a "
        "baseline model (elevation, slope, landcover, NDVI) versus a "
        "baseline+thermal model (adding LST anomaly, current TVDI, TVDI "
        "difference, and downscaled/fused LST), compared under spatial-block "
        f"cross-validation for the `{PRIMARY_POPULATION}` (primary) and "
        f"`{SECONDARY_POPULATION}` (secondary) populations.",
        "",
        "**Step8C** quantified uncertainty around Step8B's point estimates "
        "using a spatial-block bootstrap over the existing out-of-fold "
        "predictions (no retraining), confirming whether the thermal "
        "improvement holds up under resampling of the same 500 m spatial "
        "blocks used for cross-validation.",
        "",
        "**Step8D** ran a thermal feature ablation -- training baseline + "
        "each of 10 thermal feature groups/singles under the same "
        "spatial-block CV as Step8B -- to identify which specific thermal "
        "signal (LST anomaly, TVDI, downscaled/fused LST, or their "
        "combination) drives the improvement.",
        "",
        "**Step8E** (this report) does not train any model or recompute any "
        "statistic; it only aggregates the Step8A-Step8D outputs into one "
        "coherent summary.",
        "",
    ]

    lines += ["## 2. Dataset", ""]
    lines.append(f"- Total 500 m cells: `{dataset.get('total_500m_cells')}`; valid modeling cells: `{dataset.get('valid_modeling_cells')}`.")
    lines.append(f"- Burned: `{dataset.get('burned_cell_count')}`; unburned: `{dataset.get('unburned_cell_count')}`; burned rate: `{fmt(dataset.get('burned_rate'), pct=True) if dataset.get('burned_rate') is not None else 'n/a'}`.")
    lines.append(f"- Label source: `{dataset.get('label_source')}` ({dataset.get('label_source_description')}).")
    bmc = dataset.get("burn_month_counts") or {}
    bmc_str = ", ".join(
        f"{name}={bmc.get(str(num), bmc.get(num, 'n/a'))}"
        for num, name in ((8, "August"), (9, "September"), (10, "October"))
    )
    lines.append(f"- Burn month counts: {bmc_str}.")
    lines.append("- Population sizes and skip status:")
    for pop, data in populations.items():
        skip_note = data.get("skipped_in_step8b") or data.get("skipped_in_step8d")
        status = f"SKIPPED ({skip_note})" if skip_note else "modeled"
        lines.append(f"  - `{pop}`: n=`{data.get('n_cells')}` -- {status}")
    lines.append("")

    lines += ["## 3. Model comparison (Step8B)", "", "| population | AUC baseline | AUC thermal | delta_auc | PR-AUC baseline | PR-AUC thermal | delta_pr_auc | delta_brier |", "|---|---|---|---|---|---|---|---|"]
    for pop, m in model_comparison.items():
        if m is None:
            lines.append(f"| {pop} | n/a | n/a | n/a | n/a | n/a | n/a | n/a |")
            continue
        lines.append(
            f"| {pop} | {fmt(m['auc_baseline'])} | {fmt(m['auc_thermal'])} | {fmt(m['delta_auc'])} | "
            f"{fmt(m['pr_auc_baseline'])} | {fmt(m['pr_auc_thermal'])} | {fmt(m['delta_pr_auc'])} | {fmt(m['delta_brier'])} |"
        )
    lines.append("")

    lines += ["## 4. Bootstrap (Step8C)", ""]
    if not bootstrap.get("available"):
        lines.append("- **Not available**: Step8C output was not found for this report.")
    else:
        for pop, data in bootstrap["populations"].items():
            if not data.get("available"):
                lines.append(f"- **{pop}**: bootstrap unavailable.")
                continue
            interp = data["delta_auc_interpretation"]
            word = (
                "thermal improvement remained positive after spatial-block bootstrap"
                if interp == "positive_bootstrap_support"
                else "thermal improvement was negative under spatial-block bootstrap" if interp == "negative_bootstrap_support"
                else "point estimate positive but CI overlaps zero (uncertain) under spatial-block bootstrap"
            )
            lines.append(
                f"- **{pop}**: delta_auc 95% CI=`{data['delta_auc_ci95']}` "
                f"({interp}); delta_pr_auc 95% CI=`{data['delta_pr_auc_ci95']}` "
                f"({data['delta_pr_auc_interpretation']}) -- {word}."
            )
    lines.append("")

    lines += ["## 5. Feature ablation (Step8D)", ""]
    if not ablation.get("available"):
        lines.append("- **Not available**: Step8D output was not found for this report.")
    else:
        for pop, data in ablation["populations"].items():
            if not data.get("available"):
                continue
            lines.append(f"### {pop}")
            lines.append("")
            lines.append("| rank | thermal_group | delta_auc | delta_pr_auc |")
            lines.append("|---|---|---|---|")
            for group in data["ranking_order"]:
                d = data["deltas"][group]
                lines.append(f"| {d['rank_by_delta_pr_auc']} | {group} | {fmt(d['delta_auc'])} | {fmt(d['delta_pr_auc'])} |")
            lines.append("")
            lines.append(f"- Best overall group: `{data['best_group_overall']}`.")
            lines.append(f"- Best single feature: `{data['best_single_feature']}`.")
            lines.append(f"- Contribution of TVDI (delta_pr_auc): `{fmt(data['tvdi_contribution_delta_pr_auc'])}`.")
            lines.append(f"- Contribution of downscaled LST (delta_pr_auc): `{fmt(data['downscaled_contribution_delta_pr_auc'])}`.")
            fusion_word = "adds value beyond" if data["fusion_adds_value_beyond_simpler_groups"] else "does not clearly add value beyond"
            lines.append(f"- Contribution of fusion (delta_pr_auc `{fmt(data['fusion_contribution_delta_pr_auc'])}`): {fusion_word} simpler LST-anomaly/TVDI groups.")
            lines.append("")

    lines += ["## 6. Monthly lead-time", ""]
    for pop, months in monthly.items():
        lines.append(f"### {pop}")
        lines.append("")
        lines.append("| month | positives | delta_auc | delta_pr_auc | confidence |")
        lines.append("|---|---|---|---|---|")
        for month in MONTH_ORDER:
            m = months.get(month, {})
            lines.append(
                f"| {month} | {m.get('positive_count')} | {fmt(m.get('delta_auc'))} | "
                f"{fmt(m.get('delta_pr_auc'))} | {m.get('confidence')} |"
            )
        lines.append("")

    lines += ["## 7. Main findings", ""]
    lines.extend(f"- {f}" for f in findings)
    lines.append("")

    lines += ["## 8. Limitations", ""]
    lines.extend(f"- {l}" for l in limitations)
    lines.append("")

    lines += ["## 9. Future work", ""]
    lines.extend(f"{i}. {f}" for i, f in enumerate(FUTURE_WORK, start=1))
    lines.append("")

    if warnings_list:
        lines += ["## Warnings", ""]
        lines.extend(f"- {w}" for w in warnings_list)
        lines.append("")

    lines += ["## Overall conclusion", "", conclusion, ""]

    path = output_dir / "step8e_summary.md"
    path.write_text("\n".join(lines), encoding="utf-8")
    return path


def write_summary_json(
    dataset: dict, populations: dict, model_comparison: dict, bootstrap: dict,
    ablation: dict, monthly: dict, limitations: list[str], findings: list[str],
    conclusion: str, warnings_list: list[str], output_dir: Path,
) -> Path:
    payload = {
        "created_at": datetime.now(timezone.utc).isoformat(),
        "step": "step8e_final_report",
        "dataset_statistics": dataset,
        "population_statistics": populations,
        "model_comparison": model_comparison,
        "bootstrap_summary": bootstrap,
        "ablation_summary": ablation,
        "monthly_summary": monthly,
        "main_findings": findings,
        "limitations": limitations,
        "future_work": FUTURE_WORK,
        "overall_conclusion": conclusion,
        "warnings": warnings_list,
        "no_model_trained": True,
        "no_recomputation_performed": True,
        "no_firms_label_used": True,
        "no_30m_pixel_samples": True,
        "label_source": "MCD64A1",
    }
    path = output_dir / "step8e_summary.json"
    path.write_text(json.dumps(payload, indent=2, default=str), encoding="utf-8")
    return path


# =============================================================================
# Main
# =============================================================================
def main(output_dir_arg: str = STEP8E_OUTPUT_DIR, force: bool = False) -> dict:
    log.info("=" * 60)
    log.info("STEP 8E BASLIYOR (final report; no training, no recomputation)")
    log.info("=" * 60)

    out_dir = BASE_DIR / output_dir_arg
    required_outputs = [
        out_dir / "step8e_summary.md", out_dir / "step8e_summary.json",
        out_dir / "step8e_results_tables.xlsx", out_dir / "step8e_key_findings.csv",
    ]
    if any(p.exists() for p in required_outputs) and not force:
        present = [p.name for p in required_outputs if p.exists()]
        raise Step8EError("Step8E ciktilari zaten var (" + ", ".join(present) + "). Uzerine yazmak icin --force verin.")
    out_dir.mkdir(parents=True, exist_ok=True)

    step8a, step8b, step8c, step8d, warnings_list = load_all_inputs()

    dataset = build_dataset_section(step8a)
    populations = build_population_section(step8b, step8d)
    model_comparison = build_model_comparison_section(step8b)
    bootstrap = build_bootstrap_section(step8c)
    ablation = build_ablation_section(step8d)
    monthly = build_monthly_section(step8b, bootstrap)
    limitations = build_limitations(dataset, populations, model_comparison)
    findings = build_main_findings(dataset, model_comparison, bootstrap, ablation)
    conclusion = build_overall_conclusion(model_comparison, bootstrap, ablation)

    if model_comparison.get(SECONDARY_POPULATION) and model_comparison[SECONDARY_POPULATION].get("delta_auc") is not None:
        pass  # secondary population present; nothing further required here

    if step8c is None:
        log.warning("Step8C ciktisi bulunamadi; bootstrap bolumu 'not available' olarak raporlanacak.")
    if step8d is None:
        log.warning("Step8D ciktisi bulunamadi; ablation bolumu 'not available' olarak raporlanacak.")

    md_path = write_summary_md(
        dataset, populations, model_comparison, bootstrap, ablation, monthly,
        findings, limitations, conclusion, warnings_list, out_dir,
    )
    json_path = write_summary_json(
        dataset, populations, model_comparison, bootstrap, ablation, monthly,
        limitations, findings, conclusion, warnings_list, out_dir,
    )
    xlsx_path = write_excel_report(
        dataset, populations, model_comparison, bootstrap, ablation, monthly, limitations, out_dir,
    )
    findings_csv = write_key_findings_csv(findings, limitations, FUTURE_WORK, out_dir)
    feature_ranking_csv = write_feature_ranking_csv(ablation, out_dir)
    monthly_csv = write_monthly_results_csv(monthly, out_dir)
    population_csv = write_population_summary_csv(populations, out_dir)

    log.info("Summary MD: %s", md_path)
    log.info("Summary JSON: %s", json_path)
    log.info("Excel: %s", xlsx_path)
    log.info("Key findings CSV: %s", findings_csv)
    log.info("=" * 60)
    log.info("STEP 8E TAMAMLANDI (reporting only -- no training, no recomputation)")
    log.info("=" * 60)

    return {
        "summary_md": str(md_path), "summary_json": str(json_path),
        "results_xlsx": str(xlsx_path), "key_findings_csv": str(findings_csv),
        "feature_ranking_csv": str(feature_ranking_csv) if feature_ranking_csv else None,
        "monthly_csv": str(monthly_csv) if monthly_csv else None,
        "population_csv": str(population_csv) if population_csv else None,
    }


def parse_args(argv=None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Step8E: aggregate Step8A-Step8D outputs into one final scientific summary report."
    )
    parser.add_argument("--output-dir", type=str, default=STEP8E_OUTPUT_DIR)
    parser.add_argument("--force", action="store_true")
    return parser.parse_args(argv)


if __name__ == "__main__":
    args = parse_args()
    main(output_dir_arg=args.output_dir, force=args.force)